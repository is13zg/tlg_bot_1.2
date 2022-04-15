from aiogram import types, Dispatcher
from my_filters import IsAdmin, IsGod, IsThisChat
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.dispatcher.filters import Text, IsReplyFilter
from create_bot import dp, bot
import json
import init_data
import openpyxl
from aiogram.types.message import ContentType
import create_bot
import inspect
from aiogram.types import InputFile


# Чтобы загузить обновить базу книг отправьте books_for_tlg.xlsx
# и дайте команду update

# Загружает xlsx
# @dp.message_handler(IsAdmin(), content_types=['document'])
async def handle_docs(message: types.Message):
    try:
        file_id = message.document.file_id
        if message.document.file_name == init_data.UPDATE_FILE_NAME:
            file = await bot.get_file(file_id)
            file_path = file.file_path
            await bot.download_file(file_path, init_data.UPDATE_FILE_NAME)
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# обновляет json по принятому xlxsx
# @dp.message_handler(IsAdmin(), commands=['update'], commands_prefix="!/")
async def update_json(message: types.Message):
    try:
        try:
            wb = openpyxl.load_workbook(init_data.UPDATE_FILE_NAME)
            sheet = wb.active
        except Exception as e:
            await bot.send_message(init_data.my_god, e)

        columns = ["books_name", "about", "wb_url", "ozon_url", "img_url", "categories", "age_min", "age_max",
                   "webp_url", "order"]
        t_columns = []
        for i in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=i)
            if cell.value in {None, ""}:
                break
            t_columns.append(cell.value)
        if t_columns != columns:
            await bot.send_message(init_data.my_god, "при загрузке базы книг заколовки не совавли")
            await bot.send_message(message.from_user.id, "при загрузке базы книг заколовки не совавли")
            return

        ls = []
        for i in range(2, sheet.max_row):
            nd = dict()
            if sheet.cell(row=i, column=1).value in {None, ""}:
                break
            nd["books_name"] = str(sheet.cell(row=i, column=1).value)
            nd["about"] = str(sheet.cell(row=i, column=2).value)
            nd["wb_url"] = str(sheet.cell(row=i, column=3).value)
            nd["ozon_url"] = str(sheet.cell(row=i, column=4).value)
            nd["img_url"] = str(sheet.cell(row=i, column=5).value)
            nd["categories"] = str(sheet.cell(row=i, column=6).value).split(",")
            if sheet.cell(row=i, column=1).value in {None, ""}:
                break
            nd["age_min"] = int(sheet.cell(row=i, column=7).value)
            nd["age_max"] = int(sheet.cell(row=i, column=8).value)
            nd["webp_url"] = str(sheet.cell(row=i, column=9).value)
            nd["order"] = int(sheet.cell(row=i, column=10).value)

            ls.append(nd)
        md = dict()
        md["books"] = ls

        with open("books.json", "w") as file:
            json.dump(md, file)
        with open("books.json", "r") as file:
            init_data.data = json.load(file)
        init_data.update_dict_with_book_for_each_age()
        await bot.send_message(init_data.my_god, "база книг обновлена")
        await bot.send_message(message.from_user.id, "база книг обновлена")
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# отправляет file_id по photo.
# @dp.message_handler(IsAdmin(), content_types=["photo"])
async def get_photo_file_id(msg: types.Message):
    try:
        await bot.send_message(msg.from_user.id, msg.photo[0].file_id)
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# отправляет file_id по sticker
# @dp.message_handler(IsAdmin(), content_types=["sticker"])
async def get_sticker_file_id(msg: types.Message):
    try:
        await bot.send_message(msg.from_user.id, msg.sticker.file_id)
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# обработчик команды ban
# @dp.message_handler(IsAdmin(), commands=["ban"], commands_prefix="!/")
async def cmd_ban(msg: types.Message):
    try:
        if not msg.reply_to_message:
            await msg.reply("Эта команда должна быть ответом на сообщение!")
            return
        await msg.delete()
        # print(msg.reply_to_message.from_user)
        await bot.send_message(msg.chat.id, f" @{msg.reply_to_message.from_user.first_name} banned.")
        await msg.bot.kick_chat_member(chat_id=msg.chat.id, user_id=msg.reply_to_message.from_user.id)
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# обработчик команды отсылает ответ пользователю
# @dp.message_handler( IsAdmin(),IsReplyFilter(), content_types=types.ContentTypes.TEXT)
async def reply_user_message(msg: types.Message):
    try:
        ls = []
        if msg.reply_to_message.content_type == ContentType.PHOTO:
            ls = msg.reply_to_message.caption.split("\n")
        elif msg.reply_to_message.content_type == ContentType.TEXT:
            ls = msg.reply_to_message.text.split("\n")
        user_id = ls[0][8:]
        msg_id = ls[1][7:]
        await bot.send_message(chat_id=user_id, reply_to_message_id=msg_id, text=msg.text)
    except Exception as e:
        print("ls in reply_user_message ls= ", ls)
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# обработчик команды del_bad
# @dp.message_handler(IsAdmin(), commands=["check_photo"], commands_prefix="!/")
async def check_pictures(msg: types.Message):
    try:
        img_url_ls = []
        webp_url_ls = []
        for item in init_data.data["books"]:
            if item["img_url"] in init_data.NONE_SET:
                img_url_ls.append(item["books_name"])
            else:
                try:
                    await bot.send_photo(chat_id=msg.from_user.id, photo=item["img_url"])
                except Exception as e:
                    img_url_ls.append(item["books_name"])
            if item["webp_url"] in init_data.NONE_SET:
                webp_url_ls.append(item["books_name"])
            else:
                try:
                    await bot.send_sticker(chat_id=msg.from_user.id, sticker=item["webp_url"])
                except Exception as e:
                    webp_url_ls.append(item["books_name"])
        if len(img_url_ls) > 0:
            ss = "Отсуствуют или ошибки со след обычными картинками:\n" + "\n".join(img_url_ls)
            if len(ss) > 4096:
                for x in range(0, len(ss), 4096):
                    await bot.send_message(msg.from_user.id, ss[x:x + 4096])
            else:
                await bot.send_message(msg.from_user.id, ss)
        else:
            await bot.send_message(msg.from_user.id, "C обычными картинками все ок")
        if len(webp_url_ls) > 0:
            ss = "Отсуствуют или ошибки со след стикерными картинками:\n" + "\n".join(webp_url_ls)
            if len(ss) > 4096:
                for x in range(0, len(ss), 4096):
                    await bot.send_message(msg.from_user.id, ss[x:x + 4096])
            else:
                await bot.send_message(msg.from_user.id, ss)
        else:
            await bot.send_message(msg.from_user.id, "Cо стикерными картинками все ок")
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# обработчик команды del_bad
# @dp.message_handler(IsAdmin(), commands=["check_urls"], commands_prefix="!/")
async def check_urls(msg: types.Message):
    try:
        wb_url_ls = []
        ozon_url_ls = []
        for item in init_data.data["books"]:
            if item["wb_url"] in init_data.NONE_SET:
                wb_url_ls.append(item["books_name"])
            else:
                try:
                    await bot.send_message(chat_id=msg.from_user.id, text=f'{item["books_name"]} WB:\n{item["wb_url"]}')
                except Exception as e:
                    wb_url_ls.append(item["books_name"])
            if item["ozon_url"] in init_data.NONE_SET:
                ozon_url_ls.append(item["books_name"])
            else:
                try:
                    await bot.send_message(chat_id=msg.from_user.id,
                                           text=f'{item["books_name"]} OZON:\n{item["ozon_url"]}')
                except Exception as e:
                    ozon_url_ls.append(item["books_name"])
        if len(wb_url_ls) > 0:
            ss = "Отсуствуют след ссылками WB:\n" + "\n".join(wb_url_ls)
            if len(ss) > 4096:
                for x in range(0, len(ss), 4096):
                    await bot.send_message(msg.from_user.id, ss[x:x + 4096])
            else:
                await bot.send_message(msg.from_user.id, ss)
        else:
            await bot.send_message(msg.from_user.id, "C ссылками WB все ок, бот проверяет только наличие ссылки")
        if len(ozon_url_ls) > 0:
            ss = "Отсуствуют след ссылками OZON:\n" + "\n".join(ozon_url_ls)
            if len(ss) > 4096:
                for x in range(0, len(ss), 4096):
                    await bot.send_message(msg.from_user.id, ss[x:x + 4096])
            else:
                await bot.send_message(msg.from_user.id, ss)
        else:
            await bot.send_message(msg.from_user.id, "C ссылками OZON все ок, бот проверяет только наличие ссылки")
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# обработчик команды new_bad
# @dp.message_handler(IsAdmin(), commands=["new_bad"], commands_prefix="!/")
async def new_bad_words(msg: types.Message):
    try:
        new_bad_words = set(map(str.lower, msg.text.split()[1:]))
        bad_words2 = init_data.bad_words.union(new_bad_words)
        if init_data.bad_words == bad_words2:
            await bot.send_message(msg.chat.id, "Слова НЕ добавлены.")
        else:
            init_data.bad_words = bad_words2
            new_dict = dict()
            new_dict["bad_words"] = list(init_data.bad_words)
            with open("bad_words.json", "w") as file:
                json.dump(new_dict, file)
            await bot.send_message(msg.chat.id, "Слова добавлены. Текущий список:\n" + (" ").join(init_data.bad_words))
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# обработчик команды возвращает секретный токен для добавления админа
# @dp.message_handler(IsGod(), commands=["get_token"], commands_prefix="!/")
async def get_my_secret_token(msg: types.Message):
    try:
        await msg.reply(text=f"Напиши боту одноразовый код:\n/admin_me {init_data.my_secret_token}")
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# если пользователь напишет команду с токеном, станет админом
# @dp.message_handler(commands=["admin_me"], commands_prefix="!/")
async def admin_me(msg: types.Message):
    try:
        if msg.text.split()[1] == init_data.my_secret_token:
            # это фикция это до перезагрузки бота
            init_data.db.add_admin(msg.from_user.id,
                                   f"{msg.from_user.first_name} {msg.from_user.last_name} {msg.from_user.username}")
            init_data.my_secret_token = init_data.gen_my_secret_token()
            await bot.send_message(init_data.my_god,
                                   f"Новый админ {msg.from_user.first_name} {msg.from_user.last_name} {msg.from_user.username}")
            init_data.update_admins()
            await msg.reply("You are in admins")
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# обработчик команды del_bad
# @dp.message_handler(IsAdmin(), commands=["del_bad"], commands_prefix="!/")
async def del_bad_words(msg: types.Message):
    try:
        new_bad_words = set(map(str.lower, msg.text.split()[1:]))
        new_bad_words = init_data.bad_words.difference(new_bad_words)

        if (new_bad_words == init_data.bad_words):
            await bot.send_message(msg.chat.id, "Ничего не удалено, в списке не было тех слов что вы написали")
            return
        else:
            init_data.bad_words = new_bad_words
            new_dict = dict()
            new_dict["bad_words"] = list(init_data.bad_words)
            with open("bad_words.json", "w") as file:
                json.dump(new_dict, file)

            await bot.send_message(msg.chat.id, "Слова удалены.Текущий список:\n" + (" ").join(init_data.bad_words))
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# отправит сообщение всем пользователям в базе сразу после команды
# @dp.message_handler(IsGod(), commands=["send_spam"], commands_prefix="!/")
async def send_spam(msg: types.Message):
    try:
        text = msg.text[11:]
        users = init_data.db.get_users()
        for user in users:
            try:
                await bot.send_message(user[0], text)
                if int(user[1]) != 1:
                    init_data.db.set_active(user[0], 1)
            except:
                init_data.db.set_active(user[0], 0)
        await bot.send_message(init_data.my_god, "Рассылка ушла")
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# @dp.message_handler(IsGod(), commands=["del_admin"], commands_prefix="!/")
async def del_admin(msg: types.Message):
    try:
        admins = init_data.db.get_admins()
        temp_inline_markup = InlineKeyboardMarkup()
        admins_list = ""
        for i, admin in enumerate(admins):
            admins_list += f"[{str(i + 1)}] id={admin[0]} name={admin[1]}\n"
            temp_inline_btn = InlineKeyboardButton(text=str(i + 1),
                                                   callback_data=f"delete_admin_{admin[0]}")
            temp_inline_markup.insert(temp_inline_btn)

        await msg.reply("Выберите номер админа которого нужно удалить\n" + admins_list, reply_markup=temp_inline_markup)
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# @dp.callback_query_handler(Text(startswith="delete_admin_"))
async def process_callback_delete_admin(callback_query: types.CallbackQuery):
    try:
        admin_id = int(callback_query.data[13:])
        init_data.db.del_admin(admin_id)
        init_data.update_admins()
        await bot.send_message(callback_query.from_user.id, text=f"Админ {admin_id} удален")
        await callback_query.message.delete()
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# @dp.message_handler(IsAdmin(), commands=["help_admin"], commands_prefix="!/")
async def help_a(msg: types.Message):
    try:
        await msg.reply(
            """
    1. Чтобы обновить базу отправьте books_for_tlg.xlsx и введите команду /update
    2. Если отправить стикер получите его файл ИД
    3. Если отправить картинку получите его файл ИД
    4. По команде /new_bad можно добавить одно или несколько стоп-слов через пробел, данные слова нельзя будет использовать обычным пользователям в чате
    5. По команде /del_bad можно удалить одно или несколько стоп-слов через пробел
    6. По команде /check_photo бот отправит вам все фото и в конце скажет если фото которых нет или они не отправляются
    7. По команде /check_urls бот отправит вам все ссылки и в конце скажет отправит названия книг у которых нет  ссылок
    8. По команде /admin_me ТОКЕН, можно стать администратором, ТОКЕН можно получить только у Искандера
    9. Если в чате ответить на сообщение пользователя командой /ban он будет исключён из чата
    10. По команде /how_many_users выдаст количество пользователей в базе
    11. Сколько каждую книги посмотрели /view_book_stat
    12. Получить файлы данных(лог, бд, json, таблицу) /make_reserv_data
    
    Команды суперадмина
    1. /get_token - получить токен для добавления админа
    2. /del_admin -получить список админов и выбрать кого удалить
    3. /send_spam - бот разошлет сообщение после команды send_spam всем пользователям в базе
            """
        )
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# dp.register_message_handler(view_book_stat, IsAdmin(), commands=["view_book_stat"], commands_prefix="!/")
async def view_book_stat(msg: types.Message):
    try:
        book_stat = init_data.db.get_books_stat()
        text = ""

        for i, row in enumerate(sorted(book_stat, key=lambda tup: tup[1], reverse=True)):
            text += f"{i + 1}. {row[0]} - [ {row[1]} ]\n"

        text = "Просмотры каждой книги с момента запуска на сегодня:\n" + text
        if len(text) > 4096:
            for x in range(0, len(text), 4096):
                await bot.send_message(msg.from_user.id, text[x:x + 4096])
        else:
            await bot.send_message(msg.from_user.id, text)
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


# dp.register_message_handler(how_many_users, IsAdmin(), commands=["how_many_users"], commands_prefix="!/")
async def how_many_users(msg: types.Message):
    try:
        users_count = init_data.db.count_users()[0][0]
        active_users_count = init_data.db.count_active_user()[0][0]
        await msg.reply(f"Всего пользователей в базе: {users_count} из них активных: {active_users_count}")
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


#@dp.message_handler(IsAdmin(), chat_type=types.ChatType.PRIVATE,commands=["make_reserv_data"],commands_prefix="!/")
async def make_reserv_data(msg: types.Message):
    try:
        await bot.send_document(msg.chat.id,InputFile("tlg_bot_database.db"))
        await bot.send_document(msg.chat.id,InputFile("books.json"))
        await bot.send_document(msg.chat.id,InputFile("bad_words.json"))
        await bot.send_document(msg.chat.id,InputFile("mylog.log"))
        await bot.send_document(msg.chat.id, InputFile("books_for_tlg.xlsx"))
    except Exception as e:
        await create_bot.send_debug_message(__name__, inspect.currentframe().f_code.co_name, e)


def register_handlers_admin(db: Dispatcher):
    dp.register_message_handler(make_reserv_data, IsAdmin(), chat_type=types.ChatType.PRIVATE,commands=["make_reserv_data"],
                                commands_prefix="!/")
    dp.register_message_handler(handle_docs, IsAdmin(), chat_type=types.ChatType.PRIVATE, content_types=['document'])
    dp.register_message_handler(update_json, IsAdmin(), chat_type=types.ChatType.PRIVATE, commands=['update'],
                                commands_prefix="!/")
    dp.register_message_handler(get_photo_file_id, IsAdmin(), chat_type=types.ChatType.PRIVATE, content_types=["photo"])
    dp.register_message_handler(get_sticker_file_id, IsAdmin(), chat_type=types.ChatType.PRIVATE,
                                content_types=["sticker"])
    dp.register_message_handler(cmd_ban, IsAdmin(), commands=["ban"], commands_prefix="!/")
    dp.register_message_handler(check_pictures, IsAdmin(), chat_type=types.ChatType.PRIVATE, commands=["check_photo"],
                                commands_prefix="!/")
    dp.register_message_handler(check_urls, IsAdmin(), chat_type=types.ChatType.PRIVATE, commands=["check_urls"],
                                commands_prefix="!/")
    dp.register_message_handler(new_bad_words, IsAdmin(), chat_type=types.ChatType.PRIVATE, commands=["new_bad"],
                                commands_prefix="!/")
    dp.register_message_handler(del_bad_words, IsAdmin(), chat_type=types.ChatType.PRIVATE, commands=["del_bad"],
                                commands_prefix="!/")
    dp.register_message_handler(get_my_secret_token, IsGod(), chat_type=types.ChatType.PRIVATE, commands=["get_token"],
                                commands_prefix="!/")
    dp.register_message_handler(admin_me, chat_type=types.ChatType.PRIVATE, commands=["admin_me"], commands_prefix="!/")
    dp.register_message_handler(send_spam, IsGod(), chat_type=types.ChatType.PRIVATE, commands=["send_spam"],
                                commands_prefix="!/")
    dp.register_message_handler(del_admin, IsGod(), chat_type=types.ChatType.PRIVATE, commands=["del_admin"],
                                commands_prefix="!/")
    dp.register_callback_query_handler(process_callback_delete_admin, Text(startswith="delete_admin_"))
    dp.register_message_handler(help_a, IsAdmin(), chat_type=types.ChatType.PRIVATE, commands=["help_a"],
                                commands_prefix="!/")
    dp.register_message_handler(how_many_users, IsAdmin(), chat_type=types.ChatType.PRIVATE,
                                commands=["how_many_users"], commands_prefix="!/")
    dp.register_message_handler(view_book_stat, IsAdmin(), chat_type=types.ChatType.PRIVATE,
                                commands=["view_book_stat"], commands_prefix="!/")
    dp.register_message_handler(reply_user_message, IsThisChat(), IsAdmin(), IsReplyFilter(True),
                                content_types=types.ContentTypes.TEXT)
